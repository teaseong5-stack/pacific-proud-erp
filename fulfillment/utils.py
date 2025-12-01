import barcode
from barcode.writer import ImageWriter
from io import BytesIO
import base64
import openpyxl
from django.http import HttpResponse
from django.utils import timezone

def generate_barcode_image(data):
    """바코드 이미지 생성 함수"""
    Code128 = barcode.get_barcode_class('code128')
    writer = ImageWriter()
    rv = BytesIO()
    Code128(data, writer=writer).write(rv, options={'module_height': 8, 'font_size': 10})
    image_data = base64.b64encode(rv.getvalue()).decode('utf-8')
    return f"data:image/png;base64,{image_data}"

def export_to_excel(queryset, filename, columns):
    """
    엑셀 다운로드 공통 함수
    :param queryset: DB 데이터
    :param filename: 파일명
    :param columns: [(헤더, 필드명), ...]
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # 파일명 설정 (한글 깨짐 방지 등은 브라우저마다 다르나 기본형 사용)
    file_name = f"{filename}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # 1. 헤더 쓰기
    row_num = 1
    for col_num, (header, field) in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)

    # 2. 데이터 쓰기
    for obj in queryset:
        row_num += 1
        for col_num, (header, field_name) in enumerate(columns, 1):
            # 필드값 가져오기 (관계형 데이터 처리 포함)
            value = obj
            for attr in field_name.split('__'):
                if hasattr(value, attr):
                    value = getattr(value, attr)
                    if callable(value): # 메서드면 실행
                        value = value()
                else:
                    value = ""
                    break
            
            # 날짜 형식 변환
            if hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d')
            
            ws.cell(row=row_num, column=col_num).value = str(value)

    wb.save(response)
    return response